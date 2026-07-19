#!/usr/bin/env python3
"""
bls_rates.py — Build a defensible, public-domain rate dataset for Paypr's
location SEO pages (Phase 3 of SEO_PLAN.md).

Source: BLS Occupational Employment & Wage Statistics (OEWS), metro-area file.
        Childcare Workers, SOC 39-9011. U.S. government work = public domain,
        so we may republish the figures as our own presentation. We still cite
        BLS on-page (SEO_PLAN.md guardrail: never fabricate, always cite).

What it does:
  1. Downloads the OEWS metro-area workbook (oesm<YY>ma.zip) from bls.gov.
  2. Filters to the target occupation (default 39-9011).
  3. Emits docs/data/babysitter-rates.json keyed by metro slug, with the full
     hourly percentile spread (p10/p25/median/p75/p90 + mean) per metro.

Note on interpretation: 39-9011 is a *wage-employment* figure (employed/day-care
childcare workers). Informal on-demand babysitter rates run higher. Present this
as "childcare worker wage data (BLS)", not as the date-night sitter rate.

Usage:
    python3 bls_rates.py                 # auto-detect most recent year
    python3 bls_rates.py --year 2024     # pin a release year
    python3 bls_rates.py --occ 39-9011   # override occupation code
    python3 bls_rates.py --out ../docs/data/babysitter-rates.json
"""
from __future__ import annotations

import argparse
import datetime as _dt
import io
import json
import re
import sys
import urllib.error
import urllib.request
import zipfile
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required. Run:  pip install -r requirements.txt")

BLS_URL = "https://www.bls.gov/oes/special.requests/oesm{yy}ma.zip"
# BLS blocks default urllib UA; send a real browser-ish UA + contact.
HEADERS = {
    "User-Agent": "Mozilla/5.0 (paypr-rates-bot; +https://usepaypr.com) python-urllib",
    "Accept": "application/zip,application/octet-stream,*/*",
}
DEFAULT_OCC = "39-9011"  # Childcare Workers

# BLS suppression / special flags in wage cells.
_MISSING = {"", "*", "**", "~", "N/A", "NA"}
_CAPPED = "#"  # ">= $100.00/hr" (hourly) — real but top-coded.

# Hourly columns we care about → output key. Header names are matched
# case-insensitively so year-to-year casing changes don't break us.
HOURLY_COLS = {
    "H_MEAN": "mean",
    "H_PCT10": "p10",
    "H_PCT25": "p25",
    "H_MEDIAN": "median",
    "H_PCT75": "p75",
    "H_PCT90": "p90",
}


def _download(year: int) -> bytes | None:
    """Fetch the metro zip for `year`. Returns None if that year isn't published.

    BLS serves a 200 HTML stub (not a 404) for a not-yet-released year, so we
    validate the ZIP magic bytes rather than trusting the HTTP status.
    """
    yy = f"{year % 100:02d}"
    url = BLS_URL.format(yy=yy)
    req = urllib.request.Request(url, headers=HEADERS)
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            data = resp.read()
    except urllib.error.HTTPError as e:
        if e.code in (403, 404):
            return None
        raise
    return data if data[:4] == b"PK\x03\x04" else None


def _find_year(preferred: int | None) -> tuple[int, bytes]:
    """Return (year, zip_bytes). If no year given, walk back from last year."""
    if preferred is not None:
        data = _download(preferred)
        if data is None:
            sys.exit(f"OEWS metro file for {preferred} is not available on bls.gov.")
        return preferred, data
    # OEWS for reference year Y is published ~spring of Y+1, so start at last year.
    for year in range(_dt.date.today().year - 1, 2018, -1):
        data = _download(year)
        if data is not None:
            print(f"  using OEWS reference year {year}", file=sys.stderr)
            return year, data
    sys.exit("Could not locate a recent OEWS metro file on bls.gov.")


def _open_sheet(zip_bytes: bytes):
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    # Metro workbook is the MSA_* .xlsx inside the archive.
    names = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
    msa = next((n for n in names if "msa" in n.lower()), names[0] if names else None)
    if not msa:
        sys.exit("No .xlsx found inside the BLS zip.")
    wb = openpyxl.load_workbook(io.BytesIO(zf.read(msa)), read_only=True, data_only=True)
    return wb[wb.sheetnames[0]]


def _slug(area_title: str) -> str:
    s = area_title.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def _num(cell) -> float | None:
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return round(float(cell), 2)
    text = str(cell).strip()
    if text in _MISSING:
        return None
    if text == _CAPPED:
        return 100.0  # top-coded floor; flagged via `capped` below.
    try:
        return round(float(text.replace(",", "").lstrip("$")), 2)
    except ValueError:
        return None


def build(year: int | None, occ: str) -> dict:
    year, zip_bytes = _find_year(year)
    ws = _open_sheet(zip_bytes)

    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip().upper() if h is not None else "" for h in next(rows)]
    col = {name: i for i, name in enumerate(header)}

    required = ["AREA_TITLE", "OCC_CODE"]
    for r in required:
        if r not in col:
            sys.exit(f"Expected column {r!r} not found. Headers: {header}")

    def get(row, name):
        i = col.get(name)
        return row[i] if i is not None and i < len(row) else None

    metros: dict[str, dict] = {}
    for row in rows:
        if str(get(row, "OCC_CODE")).strip() != occ:
            continue
        title = get(row, "AREA_TITLE")
        if not title:
            continue
        title = str(title).strip()
        # Skip the national roll-up row if the metro file carries one.
        if title.upper().startswith(("U.S.", "UNITED STATES")):
            continue

        wages = {out: _num(get(row, src)) for src, out in HOURLY_COLS.items()}
        if all(v is None for v in wages.values()):
            continue  # fully suppressed metro — nothing to publish.

        emp = _num(get(row, "TOT_EMP"))
        metros[_slug(title)] = {
            "area_title": title,
            "state": (str(get(row, "PRIM_STATE")).strip() if get(row, "PRIM_STATE") else None),
            "area_code": (str(get(row, "AREA")).strip() if get(row, "AREA") else None),
            "employment": int(emp) if emp is not None else None,
            "hourly": wages,
            "capped": str(get(row, "H_MEDIAN")).strip() == _CAPPED,
        }

    return {
        "meta": {
            "source": "U.S. Bureau of Labor Statistics, OEWS",
            "source_url": "https://www.bls.gov/oes/",
            "occupation_code": occ,
            "occupation_title": "Childcare Workers",
            "reference_year": year,
            "generated": _dt.date.today().isoformat(),
            "units": "USD per hour",
            "note": (
                "OEWS wage-employment figures for childcare workers. Informal "
                "on-demand babysitter rates typically run higher; label accordingly."
            ),
            "metro_count": len(metros),
        },
        "metros": dict(sorted(metros.items())),
    }


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--year", type=int, default=None, help="OEWS reference year (default: auto-detect latest)")
    ap.add_argument("--occ", default=DEFAULT_OCC, help=f"SOC occupation code (default: {DEFAULT_OCC})")
    ap.add_argument(
        "--out",
        default=str(Path(__file__).resolve().parent.parent / "docs" / "data" / "babysitter-rates.json"),
        help="output JSON path",
    )
    args = ap.parse_args()

    print("Fetching BLS OEWS metro wage data…", file=sys.stderr)
    data = build(args.year, args.occ)

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(data, indent=2) + "\n")
    print(
        f"Wrote {data['meta']['metro_count']} metros "
        f"(OEWS {data['meta']['reference_year']}) → {out}",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
